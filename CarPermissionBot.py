import xlsxwriter
import openpyxl
import datetime
import smtplib
import mimetypes

from datetime import date
from openpyxl import load_workbook

vars = dict()

with open("config.txt") as f:
    for line in f:
        eq_index = line.find('=')
        var_name = line[:eq_index].strip()
        value = (line[eq_index + 1:].strip())
        vars[var_name] = value
f.close()

with open("Signature.txt") as f:
      vars['Signature'] = f.readlines()

f.close()

## Define date offset of permission

days_offset = int(vars['days_offset'])

today = date.today()
tomorrow = today + datetime.timedelta(days= days_offset)

## Define the Excel File Master in same folder
wb = load_workbook('CarPermissionMaster.xlsx')

## Define the email service parameters

From_mail = vars['From_mail']
Login_ID = vars['Login_ID']
Password = vars['Password']
smtpserver= vars['smtpserver']

To_mail = vars['To_mail']

My_message = "Dear Sir,\n\nRequest to avail self-driven car facility on %s\n" % tomorrow.strftime("%B %d, %Y")
My_message +='\n'
for lines in vars['Signature']:
    My_message += lines

My_subject = vars['My_subject']
My_subject += ' for %s' % tomorrow.strftime("%B %d, %Y")

filename = My_subject + '.xlsx'  ## Excel file created for attachment



## Generate Sheet and edit to required date
ws=wb.get_sheet_by_name('Sheet1')
ws.cell('C3').value= tomorrow.strftime("%B %d, %Y")
wb.save(filename)


def sendemail(from_addr, to_addr,
              subject, message,
              login, password,files,
              smtpserver='smtp.gmail.com:587'):


    from os.path import basename
    from email.mime.application import MIMEApplication
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email import encoders
    from email.mime.base import MIMEBase

    msg = MIMEMultipart()
    msg['Subject'] = subject 
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg.attach(MIMEText(message))

    		
    ctype, encoding = mimetypes.guess_type(files)
    if ctype is None or encoding is not None:
        ctype = "application/octet-stream"

    maintype, subtype = ctype.split("/", 1)
    
    fp = open(files, "rb")
    attachment = MIMEBase(maintype, subtype)
    attachment.set_payload(fp.read())
    fp.close()
    encoders.encode_base64(attachment)
    attachment.add_header("Content-Disposition", "attachment", filename=files)
    attachment.add_header('Content-ID', '<{}>'.format(files))
    msg.attach(attachment)
 
    server = smtplib.SMTP(smtpserver)

    print 'Gmail Service Started'
    server.starttls()
    server.login(login,password)
    problems = server.sendmail(from_addr, to_addr, msg.as_string())
    problems
    server.quit()


sendemail(From_mail,To_mail,My_subject,My_message,Login_ID,Password,filename)

print '\n\nEmail Sent Succesfully to %s' % To_mail
print 'Permission requested for %s ' %tomorrow.strftime("%B %d, %Y")

raw_input("Press Enter to close...")
